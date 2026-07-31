"""
Brand Distribution Tracker
Streamlit app untuk tracking kenaikan/penurunan BD per Brand, per Cycle, per Minggu.

Cara menjalankan:
  pip install streamlit plotly openpyxl pandas
  streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO

# ─── PAGE CONFIG ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Brand Distribution Tracker",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
.main-header {font-size:2rem;font-weight:700;color:#1a237e;margin-bottom:0.1rem;}
.sub-header  {font-size:0.9rem;color:#666;margin-bottom:1.5rem;}
div[data-testid="metric-container"] {
    background:#f5f7ff;border-radius:10px;padding:0.7rem 1rem;
    border-left:4px solid #3949ab;
}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
BD_CYCLE_LABELS = ["Cy 13'25"] + [f"Cy {i}" for i in range(1, 14)]
BD_ALL_COLS     = [85] + list(range(86, 99))
PCT_ALL_COLS    = [100] + list(range(101, 114))

CYCLE_WEEK_COLS = {
    "Cy 13'25": [28],
    "Cy 1":  [29, 30, 31, 32],
    "Cy 2":  [33, 34, 35, 36],
    "Cy 3":  [37, 38, 39, 40],
    "Cy 4":  [41, 42, 43, 44],
    "Cy 5":  [45, 46, 47, 48],
    "Cy 6":  [49, 50, 51, 52],
    "Cy 7":  [53, 54, 55, 56],
    "Cy 8":  [57, 58, 59, 60],
    "Cy 9":  [61, 62, 63, 64],
    "Cy 10": [65, 66, 67, 68],
    "Cy 11": [69, 70, 71, 72],
    "Cy 12": [73, 74, 75, 76],
    "Cy 13": [77, 78, 79, 80],
}

SKIP_AREAS = {
    "TOTAL AREA", "TOTAL", "REGIONAL 1", "REGIONAL 2", "REGIONAL 3",
    "REGIONAL 4", "REGIONAL 5", "REGIONAL 6", "TOTAL REGIONAL",
    "NATIONAL", "NASIONAL", "",
}

# ─── DATA LOADER ──────────────────────────────────────────────────────────────
def safe_float(row, idx):
    try:
        if idx >= len(row):
            return np.nan
        v = row[idx]
        if v is None:
            return np.nan
        if isinstance(v, str) and (v.strip() in ("", "-", "0.0%") or v.startswith("=")):
            return np.nan
        return float(v)
    except Exception:
        return np.nan


def parse_sheet(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    records = []
    in_data = False
    current_area = None

    for row in rows:
        if row is None or len(row) < 3:
            continue
        if row[1] == "Area" and row[2] == "Rayon":
            in_data = True
            current_area = None
            continue
        if not in_data:
            continue

        area_raw  = row[1]
        rayon_raw = row[2]

        if rayon_raw is None:
            continue
        rayon_str = str(rayon_raw).strip()
        if not rayon_str or rayon_str.startswith("=") or rayon_str == "0":
            continue
        if isinstance(area_raw, str) and area_raw.strip().upper() in SKIP_AREAS:
            continue
        if rayon_str.upper() in ("TOTAL AREA",):
            continue
        if area_raw is not None:
            area_str = str(area_raw).strip()
            if area_str and not area_str.startswith("=") and area_str.upper() not in SKIP_AREAS:
                current_area = area_str
        if current_area is None:
            continue

        rec = {
            "Area":  current_area,
            "Rayon": rayon_str,
            "OU":    safe_float(row, 5),
        }
        for lbl, ci in zip(BD_CYCLE_LABELS, BD_ALL_COLS):
            rec[f"BD_{lbl}"] = safe_float(row, ci)
        for lbl, ci in zip(BD_CYCLE_LABELS, PCT_ALL_COLS):
            rec[f"PCT_{lbl}"] = safe_float(row, ci)
        for cy_lbl, col_list in CYCLE_WEEK_COLS.items():
            for mg_idx, ci in enumerate(col_list, start=1):
                key = f"WK_{cy_lbl}_Mg{mg_idx}" if cy_lbl != "Cy 13'25" else "WK_Cy13'25"
                rec[key] = safe_float(row, ci)

        records.append(rec)

    return pd.DataFrame(records) if records else pd.DataFrame()


@st.cache_data(show_spinner="📂 Membaca file Excel…")
def load_all_brands(file_bytes):
    wb = load_workbook(file_bytes, read_only=True, data_only=True)
    brand_data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "BY BRAND":
            continue
        try:
            df = parse_sheet(wb[sheet_name])
            if not df.empty:
                brand_data[sheet_name] = df
        except Exception:
            pass
    return brand_data


def latest_cycle_with_data(df: pd.DataFrame) -> str:
    for lbl in reversed(BD_CYCLE_LABELS[1:]):
        col = f"BD_{lbl}"
        if col in df.columns and df[col].notna().any():
            return lbl
    return BD_CYCLE_LABELS[1]


def hl_delta(val):
    """Style function untuk kolom Delta — kompatibel pandas baru & lama."""
    if isinstance(val, (int, float)) and not pd.isna(val):
        if val > 0:
            return "color:#2e7d32;font-weight:bold"
        if val < 0:
            return "color:#c62828;font-weight:bold"
    return ""


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Pengaturan")
    uploaded = st.file_uploader("Upload file Excel BD (.xlsx)", type=["xlsx"])

# ─── LOAD DATA ────────────────────────────────────────────────────────────────
if uploaded:
    raw_bytes = BytesIO(uploaded.read())
    all_data  = load_all_brands(raw_bytes)
    src_label = uploaded.name
else:
    st.markdown('<p class="main-header">📊 Brand Distribution Tracker</p>', unsafe_allow_html=True)
    st.info("⬆️ Silakan **upload file Excel Brand Distribution** melalui sidebar kiri untuk memulai.")
    st.stop()

if not all_data:
    st.error("Tidak ada data yang berhasil terbaca. Pastikan format file sesuai.")
    st.stop()

# ─── FILTER SIDEBAR ───────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.subheader("🔍 Filter")
    brand_list = sorted(all_data.keys())

    # ── MULTI-SELECT BRAND ──
    selected_brands = st.multiselect(
        "Brand (pilih satu atau lebih)",
        options=brand_list,
        default=[brand_list[0]],
        key="sel_brands"
    )
    if not selected_brands:
        st.warning("Pilih minimal 1 brand.")
        st.stop()

    # Gabungkan semua brand yang dipilih
    frames = []
    for b in selected_brands:
        tmp = all_data[b].copy()
        tmp["Brand"] = b
        frames.append(tmp)
    df_combined = pd.concat(frames, ignore_index=True)

    # Filter Area
    area_opts = ["Semua Area"] + sorted(df_combined["Area"].dropna().unique().tolist())
    sel_area  = st.selectbox("Area", area_opts)

    if sel_area != "Semua Area":
        rayon_pool = sorted(df_combined[df_combined["Area"] == sel_area]["Rayon"].dropna().unique().tolist())
    else:
        rayon_pool = sorted(df_combined["Rayon"].dropna().unique().tolist())

    rayon_opts = ["Semua Rayon"] + rayon_pool
    sel_rayon  = st.selectbox("Rayon", rayon_opts)

    st.markdown("---")
    brand_label = ", ".join(selected_brands) if len(selected_brands) <= 3 else f"{len(selected_brands)} brand dipilih"
    st.caption(f"📄 Sumber: `{src_label}`\n\n🏷️ Brand tersedia: **{len(all_data)}**")

# ── Apply filter ──
dff = df_combined.copy()
if sel_area  != "Semua Area":
    dff = dff[dff["Area"]  == sel_area]
if sel_rayon != "Semua Rayon":
    dff = dff[dff["Rayon"] == sel_rayon]

if dff.empty:
    st.warning("⚠️ Tidak ada data untuk filter yang dipilih.")
    st.stop()

multi_brand = len(selected_brands) > 1

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown('<p class="main-header">📊 Brand Distribution Tracker</p>', unsafe_allow_html=True)
st.markdown(
    f'<p class="sub-header">Brand: <b>{brand_label}</b> | '
    f'Area: <b>{sel_area}</b> | Rayon: <b>{sel_rayon}</b></p>',
    unsafe_allow_html=True
)

# ─── KPI CARDS ────────────────────────────────────────────────────────────────
cy_now  = latest_cycle_with_data(dff)
idx_now = BD_CYCLE_LABELS.index(cy_now)
cy_prev = BD_CYCLE_LABELS[idx_now - 1] if idx_now > 0 else cy_now

bd_now   = dff[f"BD_{cy_now}"].sum(min_count=1)
bd_prev  = dff[f"BD_{cy_prev}"].sum(min_count=1) if f"BD_{cy_prev}" in dff.columns else np.nan
ou_total = dff["OU"].sum(min_count=1)
delta_bd = (bd_now - bd_prev) if not (pd.isna(bd_now) or pd.isna(bd_prev)) else None
pct_now  = (bd_now / ou_total * 100) if (not pd.isna(bd_now) and ou_total > 0) else None
pct_prev_val = (bd_prev / ou_total * 100) if (not pd.isna(bd_prev) and ou_total > 0) else None
delta_pct    = (pct_now - pct_prev_val) if (pct_now is not None and pct_prev_val is not None) else None

c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    st.metric("Outlet Universe", f"{ou_total:,.0f}" if not pd.isna(ou_total) else "—")
with c2:
    st.metric(f"BD {cy_now} (outlet)", f"{bd_now:,.0f}" if not pd.isna(bd_now) else "—",
              delta=f"{delta_bd:+,.0f}" if delta_bd is not None else None)
with c3:
    st.metric(f"% BD {cy_now}", f"{pct_now:.1f}%" if pct_now is not None else "—",
              delta=f"{delta_pct:+.1f}%" if delta_pct is not None else None)
with c4:
    st.metric(f"BD {cy_prev} (prev)", f"{bd_prev:,.0f}" if not pd.isna(bd_prev) else "—")
with c5:
    if delta_bd is not None:
        icon  = "📈 Naik" if delta_bd > 0 else ("📉 Turun" if delta_bd < 0 else "➡️ Sama")
        color = "normal" if delta_bd >= 0 else "inverse"
        st.metric("Trend vs Cycle Sebelumnya", icon, delta=f"{delta_bd:+,.0f}", delta_color=color)
    else:
        st.metric("Trend", "—")

st.markdown("---")

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "🔢  Tabel per Cycle",
    "📈  Chart Tren Cycle",
    "📅  Tren Mingguan",
    "🏆  Ranking & Mover",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — TABEL PER CYCLE
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("Detail BD per Area / Rayon per Cycle")

    view = st.radio(
        "Tampilkan nilai",
        ["Absolut (Jumlah Outlet)", "% BD vs Outlet Universe"],
        horizontal=True, key="t1_view"
    )

    prefix = "BD_" if view == "Absolut (Jumlah Outlet)" else "PCT_"
    fmt    = "{:,.0f}" if view == "Absolut (Jumlah Outlet)" else "{:.2f}%"

    grp_cols = (["Brand", "Area", "Rayon"] if multi_brand else ["Area", "Rayon"]) \
               if sel_rayon == "Semua Rayon" else \
               (["Brand", "Rayon"] if multi_brand else ["Rayon"])

    cy_cols  = [f"{prefix}{lbl}" for lbl in BD_CYCLE_LABELS]
    show_df  = dff[grp_cols + cy_cols].copy()
    show_df.columns = grp_cols + BD_CYCLE_LABELS

    def color_cycles(row):
        bg = [""] * len(row)
        vals = [row.get(c, np.nan) for c in BD_CYCLE_LABELS]
        for i in range(1, len(BD_CYCLE_LABELS)):
            curr = vals[i]; prev_v = vals[i-1]
            col_pos = len(grp_cols) + i
            if pd.isna(curr) or pd.isna(prev_v) or prev_v == 0:
                bg[col_pos] = ""
            elif curr > prev_v:
                bg[col_pos] = "background-color:#c8e6c9"
            elif curr < prev_v:
                bg[col_pos] = "background-color:#ffcdd2"
        return bg

    styled = (
        show_df.style
        .apply(color_cycles, axis=1)
        .format({c: fmt for c in BD_CYCLE_LABELS}, na_rep="—")
    )
    st.dataframe(styled, use_container_width=True, height=500)

    st.markdown("---")
    st.subheader(f"📊 Naik / Turun per Rayon: **{cy_prev} → {cy_now}**")

    delta_df = dff[grp_cols + [f"BD_{cy_prev}", f"BD_{cy_now}"]].copy()
    delta_df.columns = grp_cols + [cy_prev, cy_now]
    delta_df["Delta"]  = delta_df[cy_now] - delta_df[cy_prev]
    delta_df["Status"] = delta_df["Delta"].apply(
        lambda x: "⬆️ Naik" if x > 0 else ("⬇️ Turun" if x < 0 else "➡️ Sama")
    )
    delta_df = delta_df.dropna(subset=[cy_now]).sort_values("Delta", ascending=False)

    n_naik  = (delta_df["Delta"] > 0).sum()
    n_turun = (delta_df["Delta"] < 0).sum()
    n_sama  = (delta_df["Delta"] == 0).sum()

    ca, cb, cc = st.columns(3)
    ca.metric("⬆️ Naik",  n_naik)
    cb.metric("⬇️ Turun", n_turun)
    cc.metric("➡️ Sama",  n_sama)

    styled_d = (
        delta_df.style
        .map(hl_delta, subset=["Delta"])          # ← .map() bukan .applymap()
        .format({cy_prev: "{:,.0f}", cy_now: "{:,.0f}", "Delta": "{:+,.0f}"}, na_rep="—")
    )
    st.dataframe(styled_d, use_container_width=True, height=400)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CHART TREN PER CYCLE
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader(f"Tren Absolute BD per Cycle — {brand_label}")

    # Kalau multi-brand, default group by Brand; kalau single brand group by Rayon/Area
    if multi_brand:
        chart_grp_opts = ["Brand", "Rayon", "Area"]
    else:
        chart_grp_opts = ["Rayon", "Area"]

    chart_grp = st.radio("Kelompokkan berdasarkan", chart_grp_opts, horizontal=True, key="t2_grp")

    agg = dff.groupby(chart_grp)[[f"BD_{c}" for c in BD_CYCLE_LABELS]].sum(min_count=1).reset_index()
    agg.columns = [chart_grp] + BD_CYCLE_LABELS
    colors = px.colors.qualitative.Plotly

    fig = go.Figure()
    for i, (_, row) in enumerate(agg.iterrows()):
        y = [row.get(c) for c in BD_CYCLE_LABELS]
        fig.add_trace(go.Scatter(
            x=BD_CYCLE_LABELS, y=y, mode="lines+markers",
            name=str(row[chart_grp]),
            line=dict(width=2.5, color=colors[i % len(colors)]),
            marker=dict(size=8),
            hovertemplate=f"<b>{row[chart_grp]}</b><br>%{{x}}: %{{y:,.0f}} outlet<extra></extra>",
        ))
    fig.update_layout(
        title=f"Tren BD per Cycle — {brand_label}",
        xaxis_title="Cycle", yaxis_title="Jumlah Outlet (BD)",
        hovermode="x unified", height=460,
        plot_bgcolor="#fafafa", legend_title=chart_grp,
    )
    fig.update_xaxes(showgrid=True, gridcolor="#e0e0e0")
    fig.update_yaxes(showgrid=True, gridcolor="#e0e0e0")
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("Delta BD Aggregat antar Cycle")

    agg_total = agg[BD_CYCLE_LABELS].sum(min_count=1)
    deltas    = agg_total.diff().fillna(0)
    bar_cols  = ["#2e7d32" if v >= 0 else "#c62828" for v in deltas]

    fig2 = go.Figure(go.Bar(
        x=BD_CYCLE_LABELS, y=deltas.values,
        marker_color=bar_cols,
        text=[f"{v:+,.0f}" for v in deltas.values],
        textposition="outside",
        hovertemplate="%{x}: %{y:+,.0f}<extra></extra>",
    ))
    fig2.add_hline(y=0, line_color="gray", line_dash="dot")
    fig2.update_layout(
        title="Delta BD vs Cycle Sebelumnya (total terpilih)",
        xaxis_title="Cycle", yaxis_title="Delta Outlet",
        height=370, plot_bgcolor="#fafafa", showlegend=False,
    )
    st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    st.subheader("Tren % BD per Cycle")
    pct_agg = dff.groupby(chart_grp)[[f"PCT_{c}" for c in BD_CYCLE_LABELS]].mean(numeric_only=True).reset_index()
    pct_agg.columns = [chart_grp] + BD_CYCLE_LABELS

    fig3 = go.Figure()
    for i, (_, row) in enumerate(pct_agg.iterrows()):
        y = [row.get(c) for c in BD_CYCLE_LABELS]
        fig3.add_trace(go.Scatter(
            x=BD_CYCLE_LABELS, y=y, mode="lines+markers",
            name=str(row[chart_grp]),
            line=dict(width=2, dash="dot", color=colors[i % len(colors)]),
            marker=dict(size=7),
            hovertemplate=f"<b>{row[chart_grp]}</b><br>%{{x}}: %{{y:.2f}}%<extra></extra>",
        ))
    fig3.update_layout(
        title="% BD per Cycle (rata-rata)",
        xaxis_title="Cycle", yaxis_title="% BD",
        hovermode="x unified", height=400,
        plot_bgcolor="#fafafa", legend_title=chart_grp,
    )
    st.plotly_chart(fig3, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — TREN MINGGUAN
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.subheader(f"Tren BD per Minggu — {brand_label}")

    avail_cycles = [c for c in list(CYCLE_WEEK_COLS.keys())[1:]
                    if any(f"WK_{c}_Mg{mg}" in dff.columns for mg in range(1, 5))]

    sel_cycles = st.multiselect(
        "Pilih Cycle yang ditampilkan",
        options=avail_cycles,
        default=avail_cycles[-min(3, len(avail_cycles)):],
        key="t3_cycles"
    )

    # Kalau multi-brand, group per Brand; kalau single, aggregate semua
    grp_wk = "Brand" if multi_brand else None

    week_records = []
    if grp_wk:
        for brand_name in selected_brands:
            brand_dff = dff[dff["Brand"] == brand_name]
            for cy in sel_cycles:
                for mg in range(1, 5):
                    col = f"WK_{cy}_Mg{mg}"
                    if col in brand_dff.columns:
                        val = brand_dff[col].sum(min_count=1)
                        if not pd.isna(val):
                            week_records.append({
                                "Brand": brand_name,
                                "Cycle": cy,
                                "Minggu": f"Mg {mg}",
                                "label": f"{cy} Mg{mg}",
                                "BD": val
                            })
    else:
        for cy in sel_cycles:
            for mg in range(1, 5):
                col = f"WK_{cy}_Mg{mg}"
                if col in dff.columns:
                    val = dff[col].sum(min_count=1)
                    if not pd.isna(val):
                        week_records.append({
                            "Cycle": cy,
                            "Minggu": f"Mg {mg}",
                            "label": f"{cy} Mg{mg}",
                            "BD": val
                        })

    if week_records:
        wdf = pd.DataFrame(week_records)
        color_col = "Brand" if multi_brand else "Cycle"
        fig4 = px.line(
            wdf, x="label", y="BD", color=color_col, markers=True,
            labels={"BD": "Jumlah Outlet (BD)", "label": "Minggu"},
            title="Tren BD per Minggu",
            color_discrete_sequence=px.colors.qualitative.Set1,
        )
        fig4.update_layout(height=430, plot_bgcolor="#fafafa", hovermode="x unified")
        fig4.update_xaxes(tickangle=45, showgrid=True, gridcolor="#e0e0e0")
        fig4.update_yaxes(showgrid=True, gridcolor="#e0e0e0")
        st.plotly_chart(fig4, use_container_width=True)

        # Delta dalam 1 Cycle (aggregate semua brand)
        st.markdown("---")
        st.subheader("Perubahan Mingguan dalam 1 Cycle")
        if sel_cycles:
            sel_cy_detail = st.selectbox("Pilih Cycle", sel_cycles, key="t3_cy_detail")
            # Aggregate per label (gabungkan semua brand)
            cy_wdf = (
                wdf[wdf["Cycle"] == sel_cy_detail]
                .groupby("label", sort=False)["BD"].sum()
                .reset_index()
            )
            cy_wdf["Delta"] = cy_wdf["BD"].diff()

            bar_cl = ["#2e7d32" if (pd.isna(v) or v >= 0) else "#c62828" for v in cy_wdf["Delta"]]
            fig5 = go.Figure(go.Bar(
                x=cy_wdf["label"], y=cy_wdf["Delta"],
                marker_color=bar_cl,
                text=[f"{v:+,.0f}" if not pd.isna(v) else "" for v in cy_wdf["Delta"]],
                textposition="outside",
            ))
            fig5.add_hline(y=0, line_color="gray", line_dash="dot")
            fig5.update_layout(
                title=f"Delta BD antar Minggu — {sel_cy_detail}",
                height=350, plot_bgcolor="#fafafa", showlegend=False,
            )
            st.plotly_chart(fig5, use_container_width=True)

        # Heatmap per Rayon
        st.markdown("---")
        st.subheader("Heatmap BD Mingguan per Rayon")
        if sel_cycles:
            ht_cy = st.selectbox("Cycle untuk Heatmap", sel_cycles, key="t3_hm")
            hm_cols = [f"WK_{ht_cy}_Mg{mg}" for mg in range(1, 5) if f"WK_{ht_cy}_Mg{mg}" in dff.columns]
            if hm_cols:
                hm_df = dff[["Rayon"] + hm_cols].groupby("Rayon").sum(min_count=1).reset_index()
                hm_df.columns = ["Rayon"] + [f"Mg {i+1}" for i in range(len(hm_cols))]
                z_vals = hm_df[[c for c in hm_df.columns if c != "Rayon"]].values
                fig6 = go.Figure(go.Heatmap(
                    z=z_vals,
                    x=[c for c in hm_df.columns if c != "Rayon"],
                    y=hm_df["Rayon"].tolist(),
                    colorscale="Blues",
                    hoverongaps=False,
                    text=[[f"{v:,.0f}" if not pd.isna(v) else "—" for v in r] for r in z_vals],
                    texttemplate="%{text}",
                ))
                fig6.update_layout(
                    title=f"Heatmap BD per Rayon — {ht_cy}",
                    height=max(300, len(hm_df) * 28 + 100),
                    margin=dict(l=150),
                )
                st.plotly_chart(fig6, use_container_width=True)
    else:
        st.info("Belum ada data mingguan untuk filter yang dipilih.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — RANKING & BIGGEST MOVER
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.subheader(f"🏆 Ranking & Biggest Mover — {brand_label}")

    rank_base = st.selectbox(
        "Ranking berdasarkan Cycle",
        options=BD_CYCLE_LABELS[1:],
        index=max(0, BD_CYCLE_LABELS[1:].index(cy_now)) if cy_now in BD_CYCLE_LABELS[1:] else 0,
        key="t4_cy"
    )
    rank_prev = BD_CYCLE_LABELS[max(0, BD_CYCLE_LABELS.index(rank_base) - 1)]

    if multi_brand:
        rank_grp_opts = ["Brand", "Rayon", "Area"]
    else:
        rank_grp_opts = ["Rayon", "Area"]

    rank_grp = st.radio("Kelompokkan", rank_grp_opts, horizontal=True, key="t4_grp")

    if rank_grp == "Brand":
        rdf = df_combined.groupby("Brand")[[f"BD_{rank_base}", f"BD_{rank_prev}"]].sum(min_count=1).reset_index()
        name_col = "Brand"
    elif rank_grp == "Area":
        rdf = df_combined.groupby("Area")[[f"BD_{rank_base}", f"BD_{rank_prev}"]].sum(min_count=1).reset_index()
        name_col = "Area"
    else:
        rdf = df_combined.groupby(["Area", "Rayon"])[[f"BD_{rank_base}", f"BD_{rank_prev}"]].sum(min_count=1).reset_index()
        name_col = "Rayon"

    rdf.columns = [c if c not in (f"BD_{rank_base}", f"BD_{rank_prev}") else
                   (rank_base if c == f"BD_{rank_base}" else rank_prev) for c in rdf.columns]
    rdf["Delta"]  = rdf[rank_base] - rdf[rank_prev]
    rdf["Status"] = rdf["Delta"].apply(
        lambda x: "⬆️ Naik" if x > 0 else ("⬇️ Turun" if x < 0 else "➡️ Sama")
    )
    rdf = rdf.dropna(subset=[rank_base]).sort_values(rank_base, ascending=False).reset_index(drop=True)
    rdf.index += 1

    top_n = st.slider("Tampilkan Top-N", 5, min(50, len(rdf)), min(20, len(rdf)), key="t4_topn")
    rdf_top = rdf.head(top_n)

    label_col = (rdf_top["Rayon"] + " (" + rdf_top["Area"] + ")") if name_col == "Rayon" else rdf_top[name_col]

    fig7 = go.Figure(go.Bar(
        y=label_col, x=rdf_top[rank_base],
        orientation="h", marker_color="#3949ab",
        text=rdf_top[rank_base].apply(lambda v: f"{v:,.0f}" if not pd.isna(v) else "—"),
        textposition="outside",
        hovertemplate="%{y}: %{x:,.0f}<extra></extra>",
    ))
    fig7.update_layout(
        title=f"Top {top_n} {rank_grp} — BD {rank_base}",
        xaxis_title="Jumlah Outlet (BD)",
        yaxis=dict(autorange="reversed"),
        height=max(350, top_n * 28 + 80),
        plot_bgcolor="#fafafa", showlegend=False,
        margin=dict(l=220),
    )
    st.plotly_chart(fig7, use_container_width=True)

    st.markdown("---")
    st.subheader(f"Biggest Mover: {rank_prev} → {rank_base}")

    col_up, col_dn = st.columns(2)
    top_up = rdf.dropna(subset=["Delta"]).nlargest(10, "Delta")
    top_dn = rdf.dropna(subset=["Delta"]).nsmallest(10, "Delta")

    lbl_up = (top_up["Rayon"] + " (" + top_up["Area"] + ")") if name_col == "Rayon" else top_up[name_col]
    lbl_dn = (top_dn["Rayon"] + " (" + top_dn["Area"] + ")") if name_col == "Rayon" else top_dn[name_col]

    with col_up:
        st.markdown("**📈 Kenaikan Terbesar**")
        fig8 = go.Figure(go.Bar(
            y=lbl_up, x=top_up["Delta"], orientation="h",
            marker_color="#2e7d32",
            text=[f"+{v:,.0f}" for v in top_up["Delta"]], textposition="outside",
        ))
        fig8.update_layout(height=340, plot_bgcolor="#fafafa", showlegend=False,
                           yaxis=dict(autorange="reversed"),
                           xaxis_title="Delta BD", margin=dict(l=180))
        st.plotly_chart(fig8, use_container_width=True)

    with col_dn:
        st.markdown("**📉 Penurunan Terbesar**")
        fig9 = go.Figure(go.Bar(
            y=lbl_dn, x=top_dn["Delta"], orientation="h",
            marker_color="#c62828",
            text=[f"{v:,.0f}" for v in top_dn["Delta"]], textposition="outside",
        ))
        fig9.update_layout(height=340, plot_bgcolor="#fafafa", showlegend=False,
                           yaxis=dict(autorange="reversed"),
                           xaxis_title="Delta BD", margin=dict(l=180))
        st.plotly_chart(fig9, use_container_width=True)

    st.markdown("---")
    st.subheader("Tabel Lengkap Ranking")

    styled_r = (
        rdf.style
        .map(hl_delta, subset=["Delta"])          # ← .map() bukan .applymap()
        .format({rank_base: "{:,.0f}", rank_prev: "{:,.0f}", "Delta": "{:+,.0f}"}, na_rep="—")
    )
    st.dataframe(styled_r, use_container_width=True, height=500)

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.caption(f"📊 Brand Distribution Tracker | Sumber data: {src_label}")
